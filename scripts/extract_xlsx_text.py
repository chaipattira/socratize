#!/usr/bin/env python3
"""Extract text from a .xlsx file using openpyxl, output to stdout."""
import sys
import openpyxl


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: extract_xlsx_text.py <xlsx_file>", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error reading xlsx: {e}", file=sys.stderr)
        sys.exit(1)

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(v) if v is not None else "" for v in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        if rows:
            sections.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    wb.close()

    if not sections:
        print("(no content found in workbook)", file=sys.stderr)
        sys.exit(1)

    print("\n\n".join(sections))


if __name__ == "__main__":
    main()
